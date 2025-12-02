from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from config import config

def initialize_excel():
    if not config.EXCEL_PATH.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Заказы"
        
        headers = ['Дата', 'Имя', 'Фамилия', 'Телефон', 'Username']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        wb.save(config.EXCEL_PATH)

def add_order_to_excel(order_data):
    initialize_excel()
    wb = load_workbook(config.EXCEL_PATH)
    ws = wb.active
    
    next_row = ws.max_row + 1
    
    ws.cell(row=next_row, column=1, value=order_data['date'])
    ws.cell(row=next_row, column=2, value=order_data['first_name'])
    ws.cell(row=next_row, column=3, value=order_data['last_name'])
    ws.cell(row=next_row, column=4, value=order_data['phone'])
    ws.cell(row=next_row, column=5, value=order_data.get('username', ''))
    
    col = 6
    for item in order_data['items']:
        cell = ws.cell(row=next_row, column=col)
        cell.value = item['product_name']
        if next_row == 2:
            header = ws.cell(row=1, column=col)
            header.value = 'Название товара'
            header.font = Font(bold=True, color="FFFFFF")
            header.alignment = Alignment(horizontal='center')
            header.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        col += 1
        
        cell = ws.cell(row=next_row, column=col)
        cell.value = f"{item['quantity']} шт"
        if next_row == 2:
            header = ws.cell(row=1, column=col)
            header.value = 'Количество товара'
            header.font = Font(bold=True, color="FFFFFF")
            header.alignment = Alignment(horizontal='center')
            header.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        col += 1
    
    wb.save(config.EXCEL_PATH)

def get_excel_file():
    initialize_excel()
    return config.EXCEL_PATH
